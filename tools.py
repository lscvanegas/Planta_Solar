from openpyxl import Workbook,load_workbook,cell
import numpy as np
import pandas as pd
import datetime
from os import scandir
import settings

def ls2(path):
    return [obj.name for obj in scandir(path) if obj.is_file()]

headers_list= settings.headers_list

def fix_excel(filename):

	data_orig = Workbook()
	data_fixed = Workbook()
	
	data_orig = load_workbook(filename)
	sheet_in = data_orig.get_sheet_by_name("Data")
	
	data_fixed = load_workbook("datos_fixed.xlsx")
	sheet_out = data_fixed.get_sheet_by_name("Hoja1")
	
	headers_list= settings.headers_list[1:]#['WS1', 'WS2', 'ENERGY', 'IRRAD1', 'IRRAD2', 'IRRAD3', 'IRRAD4', 'IRRAD5', 'TEMP1', 'TEMP2', 'WANG']
	
	row = 2
	sheet_out.cell(1,1, "Date Time")
	finish_flag = False
	count_nan = 0
	mes = (sheet_in.cell(2, 1).value[0])
	
	#sheet_out.cell(row, 1, sheet_in.cell(row,1).value +" "+ sheet_in.cell(row,2).value)
	while finish_flag == False:
			sheet_out.cell(row, 1, str(sheet_in.cell(row,1).value) +" "+ str(sheet_in.cell(row,2).value))
			if sheet_in.cell(row+1, 1).value[0] != mes:
				finish_flag = True
			else:
				row = row+1
	
	
	for i in range(3,14):
		row2 = 1
		finish_flag_2 = False
		while finish_flag_2 == False:
			if sheet_in.cell(row2,i).value != None: #relleno de espacios vacios
				sheet_out.cell(row2, i-1, sheet_in.cell(row2,i).value)
			else:
				if i != 5:
					sheet_out.cell(row2, i-1, sheet_in.cell((row2 - 289),i).value)
					count_nan = count_nan + 1
				else:
					sheet_out.cell(row2, i-1, sheet_in.cell((row2),i).value)
				#print("Bateo en la fila y columna", (row2-289), i)
			if row2 > 1 and sheet_in.cell(row2, 1).value[0] != mes:
				finish_flag_2 = True
			else:
				row2 = row2+1
	
	for i in range(3,14):
		sheet_out.cell(1, i-1, headers_list[i-3])
	
	#print(sheet_in.cell(2371,3).value)
	print("cant de valores faltantes ", count_nan)
	data_fixed.save("datos_fixed.xlsx")

def make_str_h():
    str_h=[]
    for i in range(24):
        if i < 10:
            str_h.append('0'+str(i))
        else:
            str_h.append(str(i))
    return str_h

str_h = make_str_h()
#print(mes.describe())
def get_day_means(fecha, mes, str_h, par):
    fecha = str(fecha)[:11]
    day_means=[]
    #obtener las medias diarias
    for i in range(len(str_h)):
        if i == 0:
            day_means.append(float(np.mean(np.array(mes.loc[fecha + str_h[i]:fecha + str_h[i+1] + ':00', par]))))
        elif i !=23:
            day_means.append(float(np.mean(np.array(mes.loc[fecha + str_h[i] + ':05':fecha + str_h[i+1] + ':00', par]))))
        else:
            day_means.append(float(np.mean(np.array(mes.loc[fecha + str_h[i], par]))))
    return day_means

def get_day_energy(fecha, mes, str_h):
    fecha = str(fecha)[:11]
    day_energy=[]
    #obtener las medias diarias
    for i in range(len(str_h)):
        if str_h[i]=='00' or str_h[i]=='23':
            day_energy.append(0.0)
        else:
            day_energy.append(float(mes.loc[fecha + str_h[i+1] + ':00', 'ENERGY']))
    return day_energy
    
def get_day_hours(fecha):
    fecha = str(fecha)[:11]
    date_h_list=[]
    for i in range(len(str_h)):
        date_h_list.append(fecha + str_h[i] + ':00')
    return date_h_list


def add_day(fecha):
    return (fecha + datetime.timedelta(days=1))

def add_data_day(fecha, mes, full_data):
    temp_day_data = []
    temp_day_data.append(get_day_hours(fecha))
    for i in range (1,len(headers_list)):
        if i != 3:
            temp_day_data.append(get_day_means(fecha, mes, str_h, headers_list[i]))
        else:
            temp_day_data.append(get_day_energy(fecha, mes, str_h))
    
    for i in range(len(str_h)):
        h_reg = []
        for j in range (len(headers_list)):
            h_reg.append(temp_day_data[j][i])
        s1 = pd.Series(h_reg, index=headers_list)
        full_data = full_data.append(s1, ignore_index = True)
    return full_data


def add_data_month(full_data, month_file_name):    
	fix_excel(month_file_name)
	mes = pd.read_excel('datos_fixed.xlsx', 'Hoja1', index_col= 0, parse_dates = True)
	#aca podria ir el tratamiento de series temporales aplicado a mes con fill
	fecha_init = mes.index.get_level_values(0)[0]
	fecha = fecha_init
	
	while True:
	    full_data = add_data_day(fecha, mes, full_data)
	    old_fecha = fecha
	    fecha = add_day(fecha)
	    if fecha.month != old_fecha.month:
	        break
	return full_data


def format_dataframe(dataframe, index):
    dataframe[index] = pd.to_datetime(dataframe[index])
    dataframe.set_index(index, inplace=True)
    dataframe = dataframe.sort_values([index])
    return dataframe

#file_name ="CL1-VSO Julio 2019.xlsx"
#fix_excel(file_name)
#
#print("Done")